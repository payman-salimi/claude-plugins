#!/usr/bin/env python3
"""Build the weekly Excel tracker for an elite-coach program.

Reads a JSON program spec and writes an .xlsx workbook with:
  - one tab per training day (exercise, target sets x reps, per-set reps / weight / RPE, notes)
  - a "Weekly Cardio" tab
  - a "Progress Log" tab the client fills in at each weekly check-in

Usage:
  python3 build_tracker.py program.json --out programs/jane/week-1-tracker.xlsx
  python3 build_tracker.py program.json --week 2 --out programs/jane/week-2-tracker.xlsx

Spec format (JSON):
{
  "client": "Jane",
  "week": 1,
  "block_weeks": 4,
  "days": [
    {
      "name": "Day 1 - Upper",
      "goal": "Build pressing strength",
      "exercises": [
        {"name": "Bench press", "sets": 4, "reps": "6", "rpe": "8", "rest": "3 min", "notes": "controlled eccentric"},
        {"name": "Bent-over row", "sets": 4, "reps": "6", "rpe": "8", "rest": "2-3 min"}
      ]
    }
  ],
  "cardio": [
    {"day": "Wednesday", "type": "Zone 2 bike", "duration": "30 min", "target": "60-70% max HR, conversational"}
  ],
  "recovery": {"sleep": "7-9 h", "protein": "1.6-2.2 g/kg", "hydration": "35 mL/kg + 500 mL per training hour"}
}

Requires openpyxl (pip install openpyxl).
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    sys.stderr.write("openpyxl is required: pip install openpyxl\n")
    sys.exit(1)

HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TARGET_FILL = PatternFill("solid", fgColor="EEF3F8")
MAX_SETS_DEFAULT = 4


def style_header(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws, min_width=8, max_width=40):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max(min_width, min(max_width, length + 2))


def safe_sheet_title(name, used):
    # Excel sheet names: max 31 chars, no []:*?/\
    cleaned = "".join(ch for ch in name if ch not in '[]:*?/\\')[:31] or "Day"
    title, n = cleaned, 2
    while title in used:
        suffix = f" ({n})"
        title = cleaned[: 31 - len(suffix)] + suffix
        n += 1
    used.add(title)
    return title


def build_day_sheet(wb, day, week, max_sets, used_titles):
    ws = wb.create_sheet(safe_sheet_title(day["name"], used_titles))
    ws["A1"] = f"{day['name']} — Week {week}"
    ws["A1"].font = Font(bold=True, size=14)
    if day.get("goal"):
        ws["A2"] = f"Goal: {day['goal']}"
        ws["A2"].font = Font(italic=True)

    headers = ["Exercise", "Target"]
    for s in range(1, max_sets + 1):
        headers += [f"Set {s} reps", f"Set {s} wt", f"Set {s} RPE"]
    headers.append("Notes")

    header_row = 4
    for col, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=h)
    style_header(ws, header_row, len(headers))
    ws.freeze_panes = ws.cell(row=header_row + 1, column=3)

    row = header_row + 1
    for ex in day.get("exercises", []):
        target = f"{ex.get('sets', '')} x {ex.get('reps', '')}".strip(" x")
        if ex.get("rpe"):
            target += f" @ RPE {ex['rpe']}"
        if ex.get("rest"):
            target += f" | rest {ex['rest']}"
        ws.cell(row=row, column=1, value=ex["name"])
        ws.cell(row=row, column=2, value=target)
        ws.cell(row=row, column=2).fill = TARGET_FILL
        ws.cell(row=row, column=len(headers), value=ex.get("notes", ""))
        row += 1

    autosize(ws)
    return ws


def build_cardio_sheet(wb, cardio, week):
    ws = wb.create_sheet("Weekly Cardio")
    ws["A1"] = f"Weekly Cardio — Week {week}"
    ws["A1"].font = Font(bold=True, size=14)
    headers = ["Day", "Type", "Planned duration", "Target (HR / pace / RPE)", "Actual duration", "Avg HR", "How it felt (1-10)", "Notes"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, len(headers))
    for i, c in enumerate(cardio, start=4):
        ws.cell(row=i, column=1, value=c.get("day", ""))
        ws.cell(row=i, column=2, value=c.get("type", ""))
        ws.cell(row=i, column=3, value=c.get("duration", ""))
        ws.cell(row=i, column=4, value=c.get("target", ""))
    autosize(ws)


def build_progress_sheet(wb, spec):
    ws = wb.create_sheet("Progress Log")
    ws["A1"] = f"Progress Log — {spec.get('client', 'Client')}"
    ws["A1"].font = Font(bold=True, size=14)
    headers = [
        "Week", "Sessions planned", "Sessions done", "Bodyweight (7-day avg)",
        "Sleep avg (h)", "Stress (1-10)", "Energy (1-10)", "New pain? (where)",
        "Top set: squat pattern", "Top set: hinge", "Top set: press", "Top set: pull",
        "Coach call (push/hold/regress/deload)", "Notes",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, len(headers))
    block_weeks = int(spec.get("block_weeks", 4))
    start_week = int(spec.get("week", 1))
    for i in range(block_weeks):
        ws.cell(row=4 + i, column=1, value=start_week + i)
        ws.cell(row=4 + i, column=2, value=len(spec.get("days", [])))

    rec = spec.get("recovery") or {}
    if rec:
        r = 4 + block_weeks + 2
        ws.cell(row=r, column=1, value="Recovery targets").font = Font(bold=True)
        for j, (k, v) in enumerate(rec.items(), start=1):
            ws.cell(row=r + j, column=1, value=k.capitalize())
            ws.cell(row=r + j, column=2, value=v)
    autosize(ws)


def build_workbook(spec, week=None, max_sets=None):
    week = week or int(spec.get("week", 1))
    if max_sets is None:
        max_sets = max(
            [int(ex.get("sets", 0)) for d in spec.get("days", []) for ex in d.get("exercises", [])] + [MAX_SETS_DEFAULT]
        )
    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    for day in spec.get("days", []):
        build_day_sheet(wb, day, week, max_sets, used)
    build_cardio_sheet(wb, spec.get("cardio", []), week)
    build_progress_sheet(wb, spec)
    return wb


def main(argv=None):
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("spec", help="Path to the program spec JSON")
    p.add_argument("--out", help="Output .xlsx path (default: programs/<client>/week-<N>-tracker.xlsx)")
    p.add_argument("--week", type=int, help="Override the week number in the spec")
    p.add_argument("--max-sets", type=int, help="Number of set columns (default: largest prescribed set count, min 4)")
    args = p.parse_args(argv)

    spec = json.loads(Path(args.spec).read_text())
    week = args.week or int(spec.get("week", 1))
    out = Path(args.out) if args.out else Path("programs") / spec.get("client", "client").lower().replace(" ", "-") / f"week-{week}-tracker.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = build_workbook(spec, week=week, max_sets=args.max_sets)
    wb.save(out)
    print(out.resolve())


if __name__ == "__main__":
    main()
